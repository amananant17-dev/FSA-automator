import streamlit as st
from google import genai
import openpyxl
import json
import os
import tempfile
from io import BytesIO

# --- PAGE SETUP ---
st.set_page_config(page_title="Multi-Year Financial Parser", layout="centered")
st.title("📊 Multi-Year Financial Parser")
st.write("Upload multiple Annual Report PDFs at once. The AI will extract the data from all of them and map it to the correct columns in your Master Excel template.")

# --- API KEY CONFIGURATION ---
api_key = st.text_input("Enter your Google Gemini API Key:", type="password")

# --- UPLOADS ---
uploaded_excel = st.file_uploader("1. Drop your Master Excel Template here", type="xlsx")
# Notice: accept_multiple_files is now True
uploaded_pdfs = st.file_uploader("2. Drop ALL Annual Report PDFs here", type="pdf", accept_multiple_files=True)

# --- DYNAMIC YEAR MATCHING ---
file_year_mapping = {}
if uploaded_pdfs:
    st.write("### 3. Assign a Fiscal Year to each PDF:")
    for pdf in uploaded_pdfs:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.write(f"📄 **{pdf.name}**")
        with col2:
            # Create a unique dropdown for every file uploaded
            year = st.selectbox(f"Year", ["2022", "2023", "2024", "2025"], key=pdf.name, label_visibility="collapsed")
            file_year_mapping[pdf.name] = year

# Exact Column mapping from your INDstocks file
column_map = {"2022": 3, "2023": 4, "2024": 5, "2025": 6}

if st.button("Process All Data & Generate Excel"):
    if not api_key:
        st.error("Please enter your Gemini API Key.")
    elif not uploaded_excel or not uploaded_pdfs:
        st.error("Please upload both the Excel template and at least one PDF.")
    else:
        # Load the Excel workbook ONCE before the loop starts
        wb = openpyxl.load_workbook(uploaded_excel)
        ws_pnl = wb['P&L']
        ws_bs = wb['Balance Sheet']
        
        # Define the exact row mappings
        pnl_mapping = {
            "Revenue from operations": 6, "Other income": 7, "Employee benefit expense": 10, 
            "Finance costs": 11, "Depreciation and amortisation": 12, "Other expenses": 13
        }
        bs_mapping = {
            "Property, plant and equipment": 7, "Trade receivables": 16, "Cash and cash equivalents": 18, 
            "Total equity": 29, "Trade payables": 35
        }

        try:
            client = genai.Client(api_key=api_key)
            
            # Create a progress tracker on the screen
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Loop through every single PDF the user uploaded
            for i, pdf in enumerate(uploaded_pdfs):
                target_year = file_year_mapping[pdf.name]
                target_column = column_map[target_year]
                
                status_text.write(f"⏳ Processing {pdf.name} for FY{target_year}...")
                
                # 1. Save PDF temporarily
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
                    temp_pdf.write(pdf.read())
                    temp_pdf_path = temp_pdf.name

                # 2. Upload to Gemini
                sample_file = client.files.upload(file=temp_pdf_path, config={'display_name': f'Annual Report {target_year}'})
                
                # 3. Prompt the AI for this specific year
                prompt = f"""
                You are an expert financial analyst. Read this Annual Report.
                Extract the financial figures for the fiscal year {target_year} and output them in STRICT JSON.
                Values MUST be in crores (₹). If a value is missing, use 0. No commas in numbers.
                
                Extract exactly these keys:
                {{
                    "PnL": {{
                        "Revenue from operations": 0.0,
                        "Other income": 0.0,
                        "Employee benefit expense": 0.0,
                        "Finance costs": 0.0,
                        "Depreciation and amortisation": 0.0,
                        "Other expenses": 0.0
                    }},
                    "BalanceSheet": {{
                        "Property, plant and equipment": 0.0,
                        "Trade receivables": 0.0,
                        "Cash and cash equivalents": 0.0,
                        "Total equity": 0.0,
                        "Trade payables": 0.0
                    }}
                }}
                """
                
                response = client.models.generate_content(
                    model='gemini-2.5-flash',
                    contents=[sample_file, prompt]
                )
                
                # 4. Clean JSON
                json_str = response.text.strip().replace('```json', '').replace('```', '')
                parsed_data = json.loads(json_str)
                
                # 5. Inject into the Master Workbook in memory
                for item, row in pnl_mapping.items():
                    if item in parsed_data.get('PnL', {}):
                        ws_pnl.cell(row=row, column=target_column).value = float(parsed_data['PnL'][item])

                for item, row in bs_mapping.items():
                    if item in parsed_data.get('BalanceSheet', {}):
                        ws_bs.cell(row=row, column=target_column).value = float(parsed_data['BalanceSheet'][item])

                # 6. Cleanup temp files for this specific PDF
                os.remove(temp_pdf_path)
                client.files.delete(name=sample_file.name)
                
                # Update progress bar
                progress_bar.progress((i + 1) / len(uploaded_pdfs))

            # --- AFTER THE LOOP FINISHES ---
            status_text.write("✅ All PDFs processed successfully! Preparing your file...")
            
            # Prepare the final master file for download
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.success("🎉 Your Master Excel file is fully updated with all selected years.")
            st.download_button(
                label="⬇️ Download Multi-Year Updated Excel File",
                data=output,
                file_name="Master_Updated_Financials.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"An error occurred: {e}")
